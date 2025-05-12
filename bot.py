
import logging
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ConversationHandler, filters, ContextTypes

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

ASK_NAME, ASK_PHONE, ASK_PRODUCT_CODE, ASK_SOURCE = range(4)
EXCEL_FILE = "orders.xlsx"

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    reply_keyboard = [['🔄 Restart']]
    markup = ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True)
    await update.message.reply_text(
        "👋 Assalomu alaykum hurmatli mijoz!\n\n"
        "Siz bu yerda mahsulotlarimizni buyurtma qilishingiz mumkin.\n\n"
        "Iltimos, ismingizni yozing:",
        reply_markup=markup
    )
    return ASK_NAME

async def ask_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in ["🔄 Restart", "🛍 Yangi mahsulot buyurtma qilish"]:
        return await start(update, context)
    context.user_data['name'] = update.message.text
    contact_button = KeyboardButton('📞 Telefon raqam yuborish', request_contact=True)
    markup = ReplyKeyboardMarkup([[contact_button]], resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text("Endi telefon raqamingizni yuboring yoki yozing:", reply_markup=markup)
    return ASK_PHONE

async def ask_product_code(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.contact:
        context.user_data['phone'] = update.message.contact.phone_number
    else:
        context.user_data['phone'] = update.message.text
    await update.message.reply_text("Mahsulot kodini kiriting:")
    return ASK_PRODUCT_CODE

async def ask_source(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['product_code'] = update.message.text
    reply_keyboard = [['Telegram', 'Instagram', 'Web-sayt']]
    markup = ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text("Bizni qayerdan topdingiz?", reply_markup=markup)
    return ASK_SOURCE

def save_to_excel(user_info, timestamp):
    if os.path.exists(EXCEL_FILE):
        df_existing = pd.read_excel(EXCEL_FILE)
    else:
        df_existing = pd.DataFrame(columns=["Sana vaqti", "Ism", "Telefon", "Mahsulot kodi", "Qayerdan topdi"])
    new_entry = {
        "Sana vaqti": timestamp,
        "Ism": user_info['name'],
        "Telefon": user_info['phone'],
        "Mahsulot kodi": user_info['product_code'],
        "Qayerdan topdi": user_info['source']
    }
    df_existing = pd.concat([df_existing, pd.DataFrame([new_entry])], ignore_index=True)
    df_existing.to_excel(EXCEL_FILE, index=False)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font
    ws.auto_filter.ref = ws.dimensions
    wb.save(EXCEL_FILE)

async def done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['source'] = update.message.text
    user_info = context.user_data
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    await update.message.reply_text(
        "✅ Buyurtmangiz qabul qilindi!\n\n"
        "Operatorimiz tez orada siz bilan bog'lanadi.\n\n"
        "📍 Bizning manzilimiz: https://yandex.uz/web-maps/-/CHV9uF1f"
    )
    msg = (
        f"📝 Yangi buyurtma!\n\n"
        f"🕒 Sana vaqti: {timestamp}\n"
        f"👤 Ism: {user_info['name']}\n"
        f"📞 Telefon: {user_info['phone']}\n"
        f"🔢 Mahsulot kodi: {user_info['product_code']}\n"
        f"🌐 Qayerdan topdi: {user_info['source']}"
    )
    admin_chat_id = "8149799251"
    await context.bot.send_message(chat_id=admin_chat_id, text=msg)
    save_to_excel(user_info, timestamp)
    keyboard = [['🛍 Yangi mahsulot buyurtma qilish']]
    markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text("Yangi mahsulot buyurtma qilish uchun pastdagi tugmani bosing 👇", reply_markup=markup)
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Buyurtma bekor qilindi.")
    return ConversationHandler.END

def main():
    application = ApplicationBuilder().token("7840788383:AAEJvFtjvsle5a4H25UGexn9GIZ1aDdPIZE").build()
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            ASK_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_phone)],
            ASK_PHONE: [
                MessageHandler(filters.CONTACT, ask_product_code),
                MessageHandler(filters.TEXT & ~filters.COMMAND, ask_product_code)
            ],
            ASK_PRODUCT_CODE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_source)],
            ASK_SOURCE: [MessageHandler(filters.TEXT & ~filters.COMMAND, done)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    application.add_handler(conv_handler)
    application.run_polling()

if __name__ == '__main__':
    main()
